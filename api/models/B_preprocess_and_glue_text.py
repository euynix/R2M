import os
import re
import pandas as pd
from nltk.corpus import stopwords
import openpyxl
from api.models import PorterStemmerModified_Marketing

data_path = os.path.dirname(__file__)

stemmer = PorterStemmerModified_Marketing.PorterStemmerModified_Marketing()
specialCharacters = '.,!?:;"|()[]{}' + "'"
pre_list = {' non-': ' non', ' pre-': ' pre', ' multi-': ' multi', ' re-': ' re', ' sub-': ' sub', ' un-': ' un',
            ' inter-': ' inter', ' intra-': ' intra', ' micro-': ' micro', ' macro-': ' macro', ' post-': ' post',
            ' non ': ' non', ' pre ': ' pre', ' multi ': ' multi', ' re ': ' re', ' sub ': ' sub', ' un ': ' un',
            ' inter ': ' inter', ' intra ': ' intra', ' micro ': ' micro', ' macro ': ' macro', ' post ': ' post'}


def remove_special_characters(text):
    for letter in specialCharacters:
        text = text.replace(letter, " ")
    return text


def stem_a_word(word):
    return stemmer.stem(word)

#
# def stem_text(text):
#     all_words = text.split()
#     stemmed_words = [stem_a_word(word) for word in all_words]
#     return ' ' + ' '.join(stemmed_words) + ' '


def build_mapping_list(mapping_file):
    W = openpyxl.load_workbook(mapping_file)
    # p = W.get_sheet_by_name(name='cleaned')
    p = W['cleaned']
    old = []
    new = []
    first_row = True
    for row in p.iter_rows():
        if first_row:
            first_row = False
            continue
        old.append(' ' + row[0].value.strip().lower() + ' ')
        new.append(' ' + str(row[1].value).strip().lower() + ' ')
    return old, new


def build_new_mapping_list(mapping_file):
    W = openpyxl.load_workbook(mapping_file)
    # p = W.get_sheet_by_name(name='cust_stem_byrow')
    p = W['cust_stem_byrow']
    join_list = p['C']

    root = []
    synom = []
    first_row = True

    for i in range(1, len(join_list) - 1):
        if first_row:
            first_row = False
            continue
        tmp = join_list[i].value.strip().lower().split(";")
        for j in range(len(tmp) - 1):
            root.append(' ' + tmp[-1].strip() + ' ')
            synom.append(' ' + tmp[j].strip() + ' ')
    return root, synom


def map_wrong_words(old, new, text):
    for i in range(len(old)):
        while old[i] in text:
            text = text.replace(old[i], new[i])
    return text


def replace_pres(text):
    global pre_list
    for old, new in pre_list.items():
        text = text.replace(old, new)
    return text


def preprocess_text(text):
    global old_list, new_list, wrong_list, right_list, synom_list, root_list, before_list, after_list, stoplist, stemmer  # MUST: check which stemmer I am using!!!!!!!
    
    text = text.splitlines()
    text = " " + " ".join([onerow.strip() for onerow in text]) + " "
    text = remove_special_characters(text)
    text = replace_pres(text)
    mapping_text = map_wrong_words(old_list, new_list, text)
    new_text = re.sub(r'[^a-zA-Z]+', ' ', re.sub(r'[0-9]+', '', mapping_text))

    all_words = map_wrong_words(old_list, new_list, new_text).split()
    stemmed_words = [stem_a_word(word) for word in all_words if word not in stoplist and len(word) > 1]
    stemmed_words_join = ' ' + ' '.join([word for word in stemmed_words if word not in stoplist]) + ' '
    cleaned_text = map_wrong_words(wrong_list, right_list, stemmed_words_join)
    cleaned_text = map_wrong_words(synom_list, root_list, cleaned_text)
    cleaned_text = map_wrong_words(before_list, after_list, cleaned_text)
    return cleaned_text


def glue_text(text):
    """
    this function glue bigram/trigram as unigram according to the csv file. e.g., "global advertising" becomes "global advertising".
    """
    global before_non_glue_list, before_glue_list, after_non_glue_list, after_glue_list, before_glue_mapping_list, after_glue_mapping_list

    glue_before = map_wrong_words(before_non_glue_list, before_glue_list, text)
    aug_text = augmentation_texts(glue_before)
    glue_after = map_wrong_words(after_non_glue_list, after_glue_list, aug_text)
    glue_done = map_wrong_words(before_glue_mapping_list, after_glue_mapping_list, glue_after)
    return glue_done


def augmentation_texts(text):
    """
    this function split overlapping marketing keywords into two separate words. e.g., "global advertising strategy" will become "global advertising advertising strategy" in the text, which eventually becomes "global advertising" and "advertising strategy", two marketing keywords in the LDA input.
    """
    global aug_list

    text = text.strip()
    for target_word in aug_list.keys():
        #        print (target_word)
        target_list = target_word.split()
        done_part = ""
        while text != "":
            tmp_partition = text.partition(target_word)
            if tmp_partition[2] != "":
                list_candidate = tmp_partition[2].split(None, 1)
                len_list_candidate = len(list_candidate)
                first_word = list_candidate[0]
                if first_word in aug_list[target_word]:
                    aug_word = target_list[1] + " " + first_word
                    done_part = done_part + " " + tmp_partition[0] + tmp_partition[1] + " " + aug_word
                    if len_list_candidate == 1:
                        text = ""
                    else:
                        text = list_candidate[1].strip()
                else:
                    done_part = done_part + " " + tmp_partition[0] + tmp_partition[1] + " " + first_word
                    if len_list_candidate == 1:
                        text = ""
                    else:
                        text = list_candidate[1].strip()
            else:
                done_part = done_part + " " + text
                text = ""
        text = done_part.strip()
    return " " + done_part.strip() + " "



stoplist = stopwords.words('english')
customized_stop1 = pd.read_excel(data_path + "/stop_words.xlsx", header=0).word.tolist()
customized_stop2 = [u'may', u'also', u'however', u'would', u'thus', u'wouldn']
stoplist.extend(customized_stop1)
stoplist.extend(customized_stop2)
stoplist.remove('through')
stoplist.remove('per')
stoplist = list(set(stoplist))


old_list, new_list = build_mapping_list(data_path + "/all_unknown_words.xlsx")
wrong_list, right_list = build_mapping_list(data_path + "/Mapping_update.xlsx")
before_list, after_list = build_mapping_list(data_path + "/Customized_Stemming.xlsx")
root_list, synom_list = build_new_mapping_list(data_path + "/list_synonym_1223.xlsx")
before_non_glue_list, before_glue_list = build_mapping_list(data_path + "/glue_permutation_step1.xlsx")
after_non_glue_list, after_glue_list = build_mapping_list(data_path + "/glue_permutation_step2.xlsx")
before_glue_mapping_list, after_glue_mapping_list = build_mapping_list(data_path + "/Glue_mapping.xlsx")
permutation_list = pd.read_excel(data_path + "/permutation_augment_list_len2.xlsx", index_col=0, na_filter=False).transpose().to_dict('list')
aug_list = {}
for k, v in permutation_list.items():
    aug_list[k] = [i for i in v if i]


# if __name__ == "__main__":
#     path = "papers/JM2016JSTOR.pdf"
#     cleaned_text = clean_up(path)
#     preprocessed_text = preprocess_text(cleaned_text)
#     glued_text = glue_text(preprocessed_text)
#     # print(glued_text)




